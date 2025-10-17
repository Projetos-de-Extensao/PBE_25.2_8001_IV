from django.shortcuts import render
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.db.models import Count, Q
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from .models import (
    TipoUsuario, Curso, Sala, Usuario, Funcionario, Aluno,
    Vaga, Turma, ParticipacaoMonitoria, Presenca, Inscricao,
    HorarioDisponivel, AgendamentoMonitoria, SubmissaoHoras
)
from .serializers import (
    TipoUsuarioSerializer, CursoSerializer, SalaSerializer,
    UsuarioSerializer, FuncionarioSerializer, AlunoSerializer,
    VagaSerializer, TurmaSerializer, ParticipacaoMonitoriaSerializer,
    PresencaSerializer, InscricaoSerializer,
    HorarioDisponivelSerializer, AgendamentoMonitoriaSerializer, SubmissaoHorasSerializer,
    UserSerializer, RegisterSerializer
)
from .repository import listar_usuarios, listar_alunos, listar_cursos, listar_funcionarios, listar_inscricoes, listar_turmas, listar_participacoes_monitoria, listar_presencas, listar_salas, listar_tipos_usuario

User = get_user_model()


# ViewSets para API REST
class TipoUsuarioViewSet(viewsets.ModelViewSet):
    queryset = TipoUsuario.objects.all()
    serializer_class = TipoUsuarioSerializer


class CursoViewSet(viewsets.ModelViewSet):
    queryset = Curso.objects.all()
    serializer_class = CursoSerializer


class SalaViewSet(viewsets.ModelViewSet):
    queryset = Sala.objects.all()
    serializer_class = SalaSerializer


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer


class FuncionarioViewSet(viewsets.ModelViewSet):
    queryset = Funcionario.objects.all()
    serializer_class = FuncionarioSerializer


class AlunoViewSet(viewsets.ModelViewSet):
    queryset = Aluno.objects.all()
    serializer_class = AlunoSerializer


class VagaViewSet(viewsets.ModelViewSet):
    queryset = Vaga.objects.all()
    serializer_class = VagaSerializer


class TurmaViewSet(viewsets.ModelViewSet):
    queryset = Turma.objects.all()
    serializer_class = TurmaSerializer


class ParticipacaoMonitoriaViewSet(viewsets.ModelViewSet):
    queryset = ParticipacaoMonitoria.objects.all()
    serializer_class = ParticipacaoMonitoriaSerializer


class PresencaViewSet(viewsets.ModelViewSet):
    queryset = Presenca.objects.all()
    serializer_class = PresencaSerializer


class InscricaoViewSet(viewsets.ModelViewSet):
    queryset = Inscricao.objects.all()
    serializer_class = InscricaoSerializer


# Novas ViewSets para o sistema de monitorias

class HorarioDisponivelViewSet(viewsets.ModelViewSet):
    queryset = HorarioDisponivel.objects.all()
    serializer_class = HorarioDisponivelSerializer
    
    @action(detail=False, methods=['get'])
    def por_monitor(self, request):
        monitor_id = request.query_params.get('monitor_id')
        if monitor_id:
            horarios = self.queryset.filter(monitor_id=monitor_id, ativo=True)
            serializer = self.get_serializer(horarios, many=True)
            return Response(serializer.data)
        return Response({"error": "monitor_id é obrigatório"}, status=400)


class AgendamentoMonitoriaViewSet(viewsets.ModelViewSet):
    queryset = AgendamentoMonitoria.objects.all()
    serializer_class = AgendamentoMonitoriaSerializer
    
    @action(detail=False, methods=['get'])
    def meus_agendamentos(self, request):
        aluno_id = request.query_params.get('aluno_id')
        if aluno_id:
            agendamentos = self.queryset.filter(aluno_id=aluno_id).order_by('-data', '-horario_inicio')
            serializer = self.get_serializer(agendamentos, many=True)
            return Response(serializer.data)
        return Response({"error": "aluno_id é obrigatório"}, status=400)
    
    @action(detail=False, methods=['get'])
    def como_monitor(self, request):
        monitor_id = request.query_params.get('monitor_id')
        if monitor_id:
            agendamentos = self.queryset.filter(monitor_id=monitor_id).order_by('-data', '-horario_inicio')
            serializer = self.get_serializer(agendamentos, many=True)
            return Response(serializer.data)
        return Response({"error": "monitor_id é obrigatório"}, status=400)
    
    @action(detail=True, methods=['post'])
    def confirmar(self, request, pk=None):
        agendamento = self.get_object()
        agendamento.status = 'confirmado'
        agendamento.save()
        serializer = self.get_serializer(agendamento)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        agendamento = self.get_object()
        agendamento.status = 'cancelado'
        agendamento.save()
        serializer = self.get_serializer(agendamento)
        return Response(serializer.data)


class SubmissaoHorasViewSet(viewsets.ModelViewSet):
    queryset = SubmissaoHoras.objects.all()
    serializer_class = SubmissaoHorasSerializer
    
    @action(detail=False, methods=['get'])
    def pendentes(self, request):
        submissoes = self.queryset.filter(status='pendente').order_by('-data_submissao')
        serializer = self.get_serializer(submissoes, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def aprovar(self, request, pk=None):
        submissao = self.get_object()
        coordenador_id = request.data.get('coordenador_id')
        
        if not coordenador_id:
            return Response({"error": "coordenador_id é obrigatório"}, status=400)
        
        try:
            coordenador = Funcionario.objects.get(id=coordenador_id)
            submissao.status = 'aprovado'
            submissao.aprovado_por = coordenador
            submissao.data_aprovacao = datetime.now()
            submissao.observacoes = request.data.get('observacoes', '')
            submissao.save()
            
            serializer = self.get_serializer(submissao)
            return Response(serializer.data)
        except Funcionario.DoesNotExist:
            return Response({"error": "Coordenador não encontrado"}, status=404)
    
    @action(detail=True, methods=['post'])
    def rejeitar(self, request, pk=None):
        submissao = self.get_object()
        coordenador_id = request.data.get('coordenador_id')
        
        if not coordenador_id:
            return Response({"error": "coordenador_id é obrigatório"}, status=400)
        
        try:
            coordenador = Funcionario.objects.get(id=coordenador_id)
            submissao.status = 'rejeitado'
            submissao.aprovado_por = coordenador
            submissao.data_aprovacao = datetime.now()
            submissao.observacoes = request.data.get('observacoes', '')
            submissao.save()
            
            serializer = self.get_serializer(submissao)
            return Response(serializer.data)
        except Funcionario.DoesNotExist:
            return Response({"error": "Coordenador não encontrado"}, status=404)


# Views de autenticação

@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'user': UserSerializer(user).data,
            'refresh': str(refresh),
            'access': str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    try:
        user = User.objects.get(username=username)
        if user.check_password(password):
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'user': UserSerializer(user).data,
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        else:
            return Response({'error': 'Credenciais inválidas'}, status=status.HTTP_401_UNAUTHORIZED)
    except User.DoesNotExist:
        return Response({'error': 'Usuário não encontrado'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_profile(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


# Views para dashboard e estatísticas

@api_view(['GET'])
def dashboard_stats(request):
    """Estatísticas gerais para o dashboard"""
    
    # Contadores gerais
    total_alunos = Aluno.objects.filter(ativo=True).count()
    total_monitores = Aluno.objects.filter(monitorias__isnull=False, ativo=True).distinct().count()
    total_turmas = Turma.objects.filter(ativo=True).count()
    total_vagas = Vaga.objects.filter(ativo=True).count()
    
    # Frequência - últimos 30 dias
    data_inicio = datetime.now() - timedelta(days=30)
    presencas_mes = Presenca.objects.filter(data__gte=data_inicio).count()
    total_aulas_mes = Presenca.objects.filter(data__gte=data_inicio).values('turma', 'data').distinct().count()
    taxa_frequencia = (presencas_mes / (total_aulas_mes * total_alunos) * 100) if total_aulas_mes > 0 else 0
    
    # Inscrições por mês (últimos 6 meses)
    inscricoes_por_mes = []
    for i in range(6):
        mes = datetime.now() - timedelta(days=30*i)
        count = Inscricao.objects.filter(
            data_inscricao__year=mes.year,
            data_inscricao__month=mes.month
        ).count()
        inscricoes_por_mes.append({
            'mes': mes.strftime('%b/%Y'),
            'total': count
        })
    
    # Presença por turma
    presenca_por_turma = []
    for turma in Turma.objects.filter(ativo=True)[:10]:
        total_presencas = Presenca.objects.filter(turma=turma, presente=True).count()
        total_faltas = Presenca.objects.filter(turma=turma, presente=False).count()
        presenca_por_turma.append({
            'turma': turma.nome,
            'presencas': total_presencas,
            'faltas': total_faltas
        })
    
    # Vagas mais populares
    vagas_populares = []
    for vaga in Vaga.objects.filter(ativo=True).annotate(num_monitores=Count('monitores'))[:5]:
        vagas_populares.append({
            'nome': vaga.nome,
            'monitores': vaga.num_monitores
        })
    
    return Response({
        'totais': {
            'alunos': total_alunos,
            'monitores': total_monitores,
            'turmas': total_turmas,
            'vagas': total_vagas,
        },
        'frequencia': {
            'taxa': round(taxa_frequencia, 2),
            'presencas_mes': presencas_mes,
        },
        'inscricoes_por_mes': inscricoes_por_mes,
        'presenca_por_turma': presenca_por_turma,
        'vagas_populares': vagas_populares,
    })


# Exportação de relatórios

@api_view(['GET'])
def exportar_relatorio(request):
    """Exporta relatório completo em Excel"""
    tipo = request.query_params.get('tipo', 'geral')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Estilização
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    if tipo == 'geral':
        ws.title = "Relatório Geral"
        
        # Cabeçalhos
        headers = ['Aluno', 'Matrícula', 'Curso', 'Turma', 'Monitor', 'Presenças', 'Faltas', 'Taxa Frequência']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dados
        row = 2
        for aluno in Aluno.objects.filter(ativo=True):
            participacoes = ParticipacaoMonitoria.objects.filter(aluno=aluno)
            for part in participacoes:
                presencas = Presenca.objects.filter(aluno=aluno, turma=part.turma, presente=True).count()
                faltas = Presenca.objects.filter(aluno=aluno, turma=part.turma, presente=False).count()
                total = presencas + faltas
                taxa = (presencas / total * 100) if total > 0 else 0
                
                ws.cell(row=row, column=1, value=aluno.nome)
                ws.cell(row=row, column=2, value=aluno.matricula)
                ws.cell(row=row, column=3, value=aluno.curso.nome)
                ws.cell(row=row, column=4, value=part.turma.nome)
                ws.cell(row=row, column=5, value=part.turma.monitor.nome)
                ws.cell(row=row, column=6, value=presencas)
                ws.cell(row=row, column=7, value=faltas)
                ws.cell(row=row, column=8, value=f"{taxa:.2f}%")
                row += 1
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar e retornar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=relatorio_{tipo}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# View original para template
def sql_view(request):
    usuarios = listar_usuarios()
    alunos = listar_alunos()
    cursos = listar_cursos()
    funcionarios = listar_funcionarios()
    inscricoes = listar_inscricoes()
    turmas = listar_turmas()
    participacoes_monitoria = listar_participacoes_monitoria()
    presencas = listar_presencas()
    salas = listar_salas()
    tipos_usuario = listar_tipos_usuario()
    
    return render(request, 'sql_template.html', {'usuarios': usuarios, 
                                                 'alunos': alunos, 
                                                'cursos': cursos, 
                                                'funcionarios': funcionarios, 
                                                'inscricoes': inscricoes, 
                                                'turmas': turmas, 
                                                'participacoes_monitoria': participacoes_monitoria, 
                                                'presencas': presencas, 
                                                'salas': salas, 
                                                'tipos_usuario': tipos_usuario})
